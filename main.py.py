# -*- coding: utf-8 -*-
"""Untitled

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1oup3bPb3CYO2ZPFL0eZYXIHr7KqhQbEC
"""


from PyPDF2 import PdfReader
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

def extract_text_from_pdf(pdf_path):

    pdf_reader = PdfReader(pdf_path)
    page_content = {}

    for indx, pdf_page in enumerate(pdf_reader.pages):
        page_content[indx + 1] = pdf_page.extract_text()

    return page_content

def write_content_to_excel(content, excel_path):
    wb = Workbook()
    ws = wb.active
    header_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    cell_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws['A1'] = 'Page'
    ws['B1'] = 'Content'
    ws['A1'].fill = header_fill
    ws['B1'].fill = header_fill
    ws['A1'].border = thin_border
    ws['B1'].border = thin_border

    for page, text in content.items():
        row = [page, text]
        ws.append(row)
    for row in ws.iter_rows(min_row=2, max_col=2, max_row=ws.max_row):
        for cell in row:
            cell.fill = cell_fill
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    wb.save(excel_path)

def main():

    pdf_path = '/content/test5.pdf'
    excel_path = 'output.xlsx'


    content = extract_text_from_pdf(pdf_path)

    write_content_to_excel(content, excel_path)
    print(f"PDF content has been written to {excel_path}")

if __name__ == "__main__":
    main()

